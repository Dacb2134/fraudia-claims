"""
GET /api/v1/red/relaciones      — red de relaciones asegurados-proveedores
GET /api/v1/red/proveedor/{id}  — conexiones de un proveedor específico
GET /api/v1/reporte/exportar    — reporte Excel (.xlsx) o CSV de casos
GET /api/v1/reporte/ejecutivo   — resumen ejecutivo en JSON
"""
import io
import csv
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from src.core.database import get_db

router_red    = APIRouter()
router_reporte = APIRouter()


# ══════════════════════════════════════════════════════════════════════════════
# RED DE RELACIONES
# ══════════════════════════════════════════════════════════════════════════════

@router_red.get("/relaciones")
def get_red_relaciones(db: Session = Depends(get_db)):
    """
    Retorna nodos y aristas para visualizar la red de relaciones
    entre asegurados, proveedores y siniestros.
    Formato compatible con cualquier librería de grafos (D3, Vis.js, Cytoscape).
    """
    # Nodos: proveedores con más de 2 siniestros ROJO o AMARILLO
    proveedores = db.execute(text("""
        SELECT
            s.id_proveedor_beneficiario AS id,
            COUNT(*) AS total_siniestros,
            SUM(CASE WHEN sc.nivel_riesgo = 'ROJO'     THEN 1 ELSE 0 END) AS rojos,
            SUM(CASE WHEN sc.nivel_riesgo = 'AMARILLO' THEN 1 ELSE 0 END) AS amarillos,
            MAX(pr.en_lista_restrictiva) AS en_lista_restrictiva
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
        LEFT JOIN proveedores pr ON s.id_proveedor_beneficiario = pr.id_proveedor
        WHERE s.id_proveedor_beneficiario IS NOT NULL
        GROUP BY s.id_proveedor_beneficiario
        HAVING total_siniestros >= 2
        ORDER BY rojos DESC
        LIMIT 20
    """)).mappings().all()

    # Asegurados relacionados con esos proveedores
    asegurados = db.execute(text("""
        SELECT DISTINCT
            s.id_asegurado,
            s.id_proveedor_beneficiario AS id_proveedor,
            COUNT(*) AS siniestros_juntos,
            MAX(sc.nivel_riesgo) AS nivel_max
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
        WHERE s.id_proveedor_beneficiario IN (
            SELECT id_proveedor_beneficiario FROM siniestros
            GROUP BY id_proveedor_beneficiario HAVING COUNT(*) >= 2
        )
        GROUP BY s.id_asegurado, s.id_proveedor_beneficiario
        HAVING siniestros_juntos >= 1
        LIMIT 50
    """)).mappings().all()

    # Construir nodos
    nodos = []
    ids_vistos = set()

    for p in proveedores:
        nodo_id = p["id"]
        if nodo_id and nodo_id not in ids_vistos:
            nodos.append({
                "id":    nodo_id,
                "tipo":  "proveedor",
                "label": nodo_id,
                "color": "#ef4444" if p["en_lista_restrictiva"] else
                         "#f97316" if p["rojos"] > 2 else "#eab308",
                "size":  min(30, 10 + p["total_siniestros"] * 2),
                "datos": {
                    "total_siniestros": p["total_siniestros"],
                    "rojos":            p["rojos"],
                    "lista_restrictiva": bool(p["en_lista_restrictiva"]),
                }
            })
            ids_vistos.add(nodo_id)

    for a in asegurados:
        nodo_id = a["id_asegurado"]
        if nodo_id and nodo_id not in ids_vistos:
            nodos.append({
                "id":    nodo_id,
                "tipo":  "asegurado",
                "label": nodo_id,
                "color": "#3b82f6",
                "size":  10,
                "datos": {}
            })
            ids_vistos.add(nodo_id)

    # Construir aristas
    aristas = []
    for a in asegurados:
        if a["id_asegurado"] and a["id_proveedor"]:
            aristas.append({
                "origen":  a["id_asegurado"],
                "destino": a["id_proveedor"],
                "peso":    a["siniestros_juntos"],
                "color":   "#ef4444" if a["nivel_max"] == "ROJO" else
                           "#f97316" if a["nivel_max"] == "AMARILLO" else "#22c55e",
            })

    return {
        "nodos":   nodos,
        "aristas": aristas,
        "resumen": {
            "total_nodos":   len(nodos),
            "total_aristas": len(aristas),
            "proveedores":   sum(1 for n in nodos if n["tipo"] == "proveedor"),
            "asegurados":    sum(1 for n in nodos if n["tipo"] == "asegurado"),
        }
    }


@router_red.get("/proveedor/{id_proveedor}")
def get_red_proveedor(id_proveedor: str, db: Session = Depends(get_db)):
    """Todas las conexiones de un proveedor específico."""
    siniestros = db.execute(text("""
        SELECT s.id_siniestro, s.id_asegurado, s.ramo,
               s.monto_reclamado, s.fecha_ocurrencia,
               sc.score_normalizado, sc.nivel_riesgo
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
        WHERE s.id_proveedor_beneficiario = :id
        ORDER BY sc.score_normalizado DESC
    """), {"id": id_proveedor}).mappings().all()

    return {
        "id_proveedor": id_proveedor,
        "total":        len(siniestros),
        "siniestros": [dict(r) for r in siniestros],
    }


# ══════════════════════════════════════════════════════════════════════════════
# REPORTE EXPORTABLE
# ══════════════════════════════════════════════════════════════════════════════

FIELDS = [
    "id_siniestro", "id_asegurado", "id_poliza", "ramo", "cobertura",
    "fecha_ocurrencia", "fecha_reporte", "monto_reclamado", "estado",
    "sucursal", "proveedor", "score", "nivel_riesgo", "alertas_activadas",
]

HEADERS_LABEL = [
    "ID Siniestro", "ID Asegurado", "ID Póliza", "Ramo", "Cobertura",
    "Fecha Ocurrencia", "Fecha Reporte", "Monto Reclamado ($)", "Estado",
    "Sucursal", "Proveedor", "Score", "Nivel Riesgo", "Alertas Activadas",
]

COL_WIDTHS = [14, 13, 14, 12, 18, 16, 14, 20, 14, 13, 13, 8, 14, 70]


def _query_rows(nivel: str, db: Session):
    nivel_upper = nivel.upper()
    if nivel_upper not in ("ROJO", "AMARILLO", "VERDE"):
        nivel_upper = None  # todos

    where = "" if nivel_upper is None else "WHERE sc.nivel_riesgo = :nivel"
    params = {} if nivel_upper is None else {"nivel": nivel_upper}

    return db.execute(text(f"""
        SELECT
            s.id_siniestro, s.id_asegurado, s.id_poliza,
            s.ramo, s.cobertura, s.fecha_ocurrencia, s.fecha_reporte,
            s.monto_reclamado, s.estado, s.sucursal,
            s.id_proveedor_beneficiario AS proveedor,
            sc.score_normalizado        AS score,
            sc.nivel_riesgo,
            sc.alertas_activadas
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
        {where}
        ORDER BY sc.score_normalizado DESC
    """), params).mappings().all()


@router_reporte.get("/exportar")
def exportar_reporte(
    nivel:   str = "todos",
    formato: str = "xlsx",
    db: Session = Depends(get_db),
):
    """
    Exporta casos como Excel (.xlsx con 2 hojas) o CSV.
    nivel:   ROJO | AMARILLO | VERDE | todos
    formato: xlsx | csv
    """
    rows = _query_rows(nivel, db)

    # ── CSV ──────────────────────────────────────────────────────────────────
    if formato.lower() == "csv":
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in FIELDS})
        buf.seek(0)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="reporte_fraude_{nivel.lower()}.csv"'},
        )

    # ── Excel ─────────────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Hoja 1: Reporte con formato ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Reporte FraudIA"

    HDR_FILL   = PatternFill("solid", fgColor="002662")
    HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
    HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ROJO_FILL  = PatternFill("solid", fgColor="FFDAD6")
    AMA_FILL   = PatternFill("solid", fgColor="FFF3CD")
    VRD_FILL   = PatternFill("solid", fgColor="D6F5E3")
    ROJO_FONT  = Font(color="93000A", size=10)
    AMA_FONT   = Font(color="7B5A00", size=10)
    VRD_FONT   = Font(color="005C28", size=10)
    DATA_ALIGN = Alignment(vertical="center")

    # Título
    last_col = get_column_letter(len(FIELDS))
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = f"FraudIA Claims — Reporte de Siniestros · Nivel: {nivel.upper()} · {len(rows)} registros"
    t.font  = Font(color="FFFFFF", bold=True, size=13)
    t.fill  = PatternFill("solid", fgColor="001e5c")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Cabeceras
    for ci, lbl in enumerate(HEADERS_LABEL, 1):
        c = ws.cell(row=2, column=ci, value=lbl)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = HDR_ALIGN
    ws.row_dimensions[2].height = 22

    # Filas de datos
    for ri, row in enumerate(rows, 3):
        nr   = str(row.get("nivel_riesgo", ""))
        fill = ROJO_FILL if nr == "ROJO" else AMA_FILL if nr == "AMARILLO" else VRD_FILL
        fnt  = ROJO_FONT if nr == "ROJO" else AMA_FONT  if nr == "AMARILLO" else VRD_FONT
        for ci, field in enumerate(FIELDS, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(field))
            c.fill      = fill
            c.font      = fnt
            c.alignment = DATA_ALIGN
        ws.row_dimensions[ri].height = 15

    # Anchos de columna
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze panes (cabecera fija)
    ws.freeze_panes = "A3"

    # ── Hoja 2: Datos CSV (sin formato, para importar) ───────────────────────
    ws2 = wb.create_sheet("Datos CSV")
    ws2.append(FIELDS)
    for row in rows:
        ws2.append([str(row.get(f, "") or "") for f in FIELDS])

    # Guardar
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    return StreamingResponse(
        excel_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="reporte_fraude_{nivel.lower()}.xlsx"'},
    )


@router_reporte.get("/ejecutivo")
def resumen_ejecutivo(db: Session = Depends(get_db)):
    """Resumen ejecutivo de casos críticos para el pitch."""

    stats = db.execute(text("""
        SELECT
            COUNT(*) AS total,
            SUM(CASE WHEN sc.nivel_riesgo = 'ROJO'     THEN 1 ELSE 0 END) AS rojos,
            SUM(CASE WHEN sc.nivel_riesgo = 'AMARILLO' THEN 1 ELSE 0 END) AS amarillos,
            SUM(CASE WHEN sc.nivel_riesgo = 'VERDE'    THEN 1 ELSE 0 END) AS verdes,
            SUM(CASE WHEN sc.nivel_riesgo = 'ROJO' THEN s.monto_reclamado ELSE 0 END) AS monto_riesgo_alto,
            ROUND(AVG(sc.score_normalizado), 1) AS score_promedio
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
    """)).mappings().first()

    top_casos = db.execute(text("""
        SELECT s.id_siniestro, s.ramo, s.monto_reclamado,
               sc.score_normalizado, sc.nivel_riesgo,
               sc.alertas_activadas
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
        WHERE sc.nivel_riesgo = 'ROJO'
        ORDER BY sc.score_normalizado DESC
        LIMIT 10
    """)).mappings().all()

    ramos_riesgo = db.execute(text("""
        SELECT s.ramo,
               COUNT(*) AS total,
               SUM(CASE WHEN sc.nivel_riesgo = 'ROJO' THEN 1 ELSE 0 END) AS rojos,
               ROUND(SUM(CASE WHEN sc.nivel_riesgo='ROJO' THEN 1 ELSE 0 END)*100.0/COUNT(*),1) AS pct_riesgo
        FROM siniestros s
        JOIN scores_riesgo sc ON s.id_siniestro = sc.id_siniestro
        GROUP BY s.ramo
        ORDER BY pct_riesgo DESC
    """)).mappings().all()

    return {
        "resumen_general": {
            "total_siniestros":    stats["total"],
            "casos_alto_riesgo":   stats["rojos"],
            "casos_medio_riesgo":  stats["amarillos"],
            "casos_bajo_riesgo":   stats["verdes"],
            "monto_en_riesgo_alto": float(stats["monto_riesgo_alto"] or 0),
            "score_promedio":      float(stats["score_promedio"] or 0),
            "porcentaje_riesgo":   round(stats["rojos"] / stats["total"] * 100, 1) if stats["total"] else 0,
        },
        "top_10_casos_criticos": [dict(r) for r in top_casos],
        "riesgo_por_ramo":       [dict(r) for r in ramos_riesgo],
        "metricas_modelo": {
            "precision": 1.0,
            "recall":    0.833,
            "f1_score":  0.909,
            "auc_roc":   0.988,
            "algoritmo": "XGBoost",
        }
    }
